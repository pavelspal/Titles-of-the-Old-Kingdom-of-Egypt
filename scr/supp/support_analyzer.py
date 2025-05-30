try:
    from supp.support_constants import *
except:
    from support_constants import *

# make statistic description for given dataframe
def describe(df):
    # df ... pandas dataset to describe
    import numpy as np
    import pandas as pd

    # compute nan percentage in columns
    def get_nan_ratio(data):
        # data ... data from which compute the ratio
        n_row = data.shape[0]
        non_nan_count = n_row - pd.isna(data).to_numpy().sum(axis=0)
        nan_ratio = np.round((1-non_nan_count/n_row) * 100, decimals=1)
        return [non_nan_count.tolist(), nan_ratio.tolist()]
    # compute description for continuous variables
    def describe_continuous(df_cont):
        if df_cont.shape[1] == 0:
            return [['None']]
        # pandas choose by itself what is continuous (numeric) variable
        statistic = df_cont.describe().transpose()
        non_nan_count, nan_ratio = get_nan_ratio(df_cont)
        duplicated = [df_cont[col].duplicated().sum() for col in df_cont.columns]
        statistic.insert(0, 'name', statistic.index)
        statistic.insert(2, 'non_nan_count', statistic['count'])
        statistic['count'] = df_cont.shape[0]
        statistic.insert(3, 'nan_ratio', nan_ratio)
        statistic.insert(4, 'duplicated', duplicated)
        statistic = round_mean_and_std(statistic)
        return [statistic.columns.tolist()] + statistic.values.tolist()
    # compute description for factorial variables
    def describe_factorial(df_fact):
        if df_fact.shape[1] == 0:
            return [['None']]
        total_counts = df_fact.shape[0]
        # df_fact ... df with factorial variable only
        non_nan_count, nan_ratio = get_nan_ratio(df_fact)
        counts = [content.value_counts() for factor_name, content in df_fact.items()]
        uniques = [[f'{count}:{factor_level}' for factor_level, count in factor.items()] for factor in counts]
        statistic = []
        for factor_name, non_nan, ratio, levels in zip(df_fact.columns.tolist(), non_nan_count, nan_ratio, uniques):
            count_vs_level = np.round(non_nan/len(levels), decimals=1)
            statistic.append([factor_name, total_counts, non_nan, ratio, len(levels), count_vs_level] + levels)
        label_fact = ['name', 'count', 'non_nan_count', 'nan_ratio', 'levels', 'non_nan/level', 'level_distribution']
        return [label_fact] + statistic

    # compute continuous variables
    df_continuous = df.select_dtypes(include=np.number)
    continuous = describe_continuous(df_continuous)

    # compute factorial variables
    # get colum names of factorial variable
    factorial_columns = [col for col in df.columns if col not in df_continuous.columns]
    factorial = describe_factorial(df[factorial_columns])

    return [continuous, factorial]


# change the width of each column in the Excel file
# with is set to max cell-width in given column
def adjust_cell_width(workbook):
    # Iterate over each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        # Access the sheet
        sheet = workbook[sheet_name]
        # Iterate through each column in the sheet
        for column_cells in sheet.columns:
            # Get the maximum length of the content in each column
            max_length = 0
            for cell in column_cells:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
            if max_length > 50:
                max_length = 50
            # Set the column width based on the maximum length
            adjusted_width = max_length + 1 # Add a little padding
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    return workbook


# round given columns in the Excel file to be appropriate long
# reason: 'mean' and 'std' had to many decimals number -> "unclear cells"
def round_mean_and_std(df, columns=['mean', 'std'], significant_num=3):
    import numpy as np
    def my_round(val, dec):
        return round(val, dec)

    values = df[columns]
    no_nan_rows = values.notnull().all(axis=1)
    values = values.loc[no_nan_rows, :]
    dec = significant_num - np.log10(np.abs(values).astype(float)).astype(int)
    v_round = np.vectorize(my_round)
    df.loc[no_nan_rows, columns] = v_round(values, dec)
    return df

def make_excel(df, path, date=None):
    import os
    import openpyxl
    import re
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    continuous, factorial = describe(df)
    # Create a new workbook
    workbook = openpyxl.Workbook()
    # Create a new sheet in the new workbook
    sheet_content = workbook.create_sheet(title='content')

    # create content sheet
    continuous_names = set([row[0] for row in continuous[1:]])
    sheet_content.append(['name', 'type'])
    for row in df.columns:
        category = 'continuous' if row in continuous_names else 'factorial'
        sheet_content.append([row] + [category])
    # add date in the last row
    if date is not None:
        sheet_content.append([date])
    # Create a new sheet for description od continuous variables
    sheet_con = workbook.create_sheet(title='continuous')
    for row in continuous:
        sheet_con.append(row)
    sheet_fact = workbook.create_sheet(title='factorial')
    try:
        end_numeric = factorial[0].index('level_distribution')
    except:
        end_numeric = 4
    for row in factorial:
        # suppress invalid characters
        new_row = row[:end_numeric] + [ILLEGAL_CHARACTERS_RE.sub(r'', str(el)) for el in row[end_numeric:]]
        sheet_fact.append(new_row)

    # Remove the default sheet (Sheet)
    default_sheet = workbook["Sheet"]
    workbook.remove(default_sheet)
    # adjust cell width
    new_wb = adjust_cell_width(workbook)
    # Save the workbook to a file
    new_wb.save(path + '.xlsx')
    print(f'File {os.path.basename(path)} saved into {path}.')


def make_excel_analysis(df, file_name, save_path=None):
    import os
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    import openpyxl
    from datetime import datetime
    try:
        from supp.support_load import get_path
    except:
        from support_load import get_path

    if save_path is None:
        save_path = PATH_DESCRIPTION

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    # Create a new sheet in the new workbook
    sheet_content = new_wb.create_sheet(title='content')

    continuous, factorial = describe(df)

    # create content sheet
    continuous_names = set([row[0] for row in continuous[1:]])
    sheet_content.append(['name', 'type'])
    for row in list(df.columns):
        category = 'continuous' if row in continuous_names else 'factorial'
        sheet_content.append([row] + [category])
    # add date in the last row
    current_date = datetime.now().strftime('%Y-%m-%d')
    sheet_content.append([current_date])
    # Create a new sheet for description od continuous variables
    sheet_con = new_wb.create_sheet(title='continuous')
    for row in continuous:
        sheet_con.append(row)
    sheet_fact = new_wb.create_sheet(title='factorial')
    try:
        end_numeric = factorial[0].index('level_distribution')
    except:
        end_numeric = 4
    for row in factorial:
        # suppress invalid characters
        new_row = row[:end_numeric] + [ILLEGAL_CHARACTERS_RE.sub(r'', str(el)) for el in row[end_numeric:]]
        sheet_fact.append(new_row)
    sheet_fact_T = new_wb.create_sheet(title='factorial_transpose')
    # transpose factorial list
    max_len_factorial = max(len(row) for row in factorial)
    for row in factorial:
        while len(row) < max_len_factorial:
            row.append('')
    factorial = list(zip(*factorial))
    for index, row in enumerate(factorial):
        # suppress invalid characters
        if index >= end_numeric:
            row = [ILLEGAL_CHARACTERS_RE.sub(r'', str(el)) for el in row]
        sheet_fact_T.append(row)

    # Remove the default sheet (Sheet)
    default_sheet = new_wb["Sheet"]
    new_wb.remove(default_sheet)
    # adjust cell width
    new_wb = adjust_cell_width(new_wb)
    # Save the workbook to a file
    path = get_path(file_name, folder=save_path)
    # hardcode format to XLSX
    path = os.path.splitext(path)[0] + '.xlsx'
    # ensure the directory exists
    os.makedirs(os.path.dirname(path), exist_ok=True)
    # save file
    new_wb.save(path)
    print(f'Excel analysis save into {path}')

