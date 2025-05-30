import os
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import openpyxl

try:
    from supp.support_analyzer import *
    from supp.support_constants import *
    from supp.support_load import get_path
except:
    from support_analyzer import *
    from support_constants import *
    from support_load import get_path


def make_excel_analysis(df, file_name, save_path=PATH_DESCRIPTION):

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    # Create a new sheet in the new workbook
    sheet_content = new_wb.create_sheet(title='content')

    continuous, factorial = describe(df)

    # create content sheet
    continuous_names = set([row[0] for row in continuous[1:]])
    sheet_content.append(['name', 'type'])
    for row in list(df.columns):
        category = 'continuous' if row[0] in continuous_names else 'factorial'
        sheet_content.append([row]+ [category])
    # Create a new sheet for description od continuous variables
    sheet_con = new_wb.create_sheet(title='continuous')
    for row in continuous:
        sheet_con.append(row)
    sheet_fact = new_wb.create_sheet(title='factorial')
    try:
        end_numeric = factorial[0].index('level_distribution')
    except:
        end_numeric = 4
    for row in factorial:
        # suppress invalid characters
        new_row = row[:end_numeric] + [ILLEGAL_CHARACTERS_RE.sub(r'', str(el)) for el in row[end_numeric:]]
        sheet_fact.append(new_row)
    sheet_fact_T = new_wb.create_sheet(title='factorial_transpose')
    # transpose factorial list
    max_len_factorial = max(len(row) for row in factorial)
    for row in factorial:
        while len(row) < max_len_factorial:
            row.append('')
    factorial = list(zip(*factorial))
    for index, row in enumerate(factorial):
        # suppress invalid characters
        if index >= end_numeric:
            row = [ILLEGAL_CHARACTERS_RE.sub(r'', str(el)) for el in row]
        sheet_fact_T.append(row)

    # Remove the default sheet (Sheet)
    default_sheet = new_wb["Sheet"]
    new_wb.remove(default_sheet)
    # adjust cell width
    new_wb = adjust_cell_width(new_wb)
    # Save the workbook to a file
    path = get_path(file_name, folder=save_path)
    # hardcode format to XLSX
    path = os.path.splitext(path)[0] + '.xlsx'
    new_wb.save(path)
    print(f'Excel analysis save into {path}')
