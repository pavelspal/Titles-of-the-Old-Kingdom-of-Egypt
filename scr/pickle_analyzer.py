import os
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import openpyxl

from supp.support_load import load_pickle, get_name_dicts
from supp.support_analyzer import describe, adjust_cell_width


# ------------------------------
# FILE BIO
# ------------------------------
# This script analyzes maatbase pickle file.
# It makes:
#       - one Excel file with all dfs in the pickle (and its variables)
#       - folder 'table' with detail description for each dataset


# path to separated sheets
save_path = r'excel_dfs_description/'
# name of concat dfs names
save_name = 'dfs_labels.xlsx'
# Check if the folder already exists
if not os.path.exists(save_path):
    # If not, create the folder
    os.makedirs(save_path[:-1])


# ------------------------------
# LOAD DATA
# ------------------------------
dfs, dfs_name, dfs_export_date = load_pickle()


# ------------------------------
# CREATE EXCELL FILE
# ------------------------------
# creates one big Excel file
#  - first sheets contains list of all dfs
#  - following sheets refer to particular df and shows list of its columns (variables)

# Create a new Workbook
workbook = openpyxl.Workbook()

# iton ... dictionary: index of a table -> the name of the table
# ntoi ... dictionary: name of a table -> index of the table
iton, ntoi = get_name_dicts(dfs_name)
# length of sheet name in Excel file (for Excel compatibility)
shortcut_size = 30
# stoo ... dictionary: short name of a table -> full name of the table
stoo = {name[:shortcut_size]:name for name in dfs_name}

# make content sheet (with list of all dfs)
sheet = workbook.create_sheet(title='content')
for key, name in iton.items():
    sheet.append([key, name])

# for each dfs make a sheet filled with column names
for key, name in iton.items():
    # length of sheet name is limited to 30 chars -> use shortcut
    sheet = workbook.create_sheet(title=name[:shortcut_size])
    columns = dfs[key].columns
    for i, col in enumerate(columns):
        sheet.append([col])

# Remove the default sheet (Sheet)
default_sheet = workbook["Sheet"]
workbook.remove(default_sheet)
# Save the workbook to a file
workbook.save(r'data/' + save_name)


# ------------------------------
# SPLIT EXCEL FILE INTO SHEETS
# ------------------------------
# creates a single Excel file for each table
# - first sheet summarizes all variables in the table
# - second sheet deeply describes all continuous variables
# - third sheet deeply describes all factorial variables

# Iterate over each sheet in the workbook
for sheet_name in workbook.sheetnames:
    print(f'sheet {sheet_name} opening')
    # Access the sheet from the original workbook
    original_sheet = workbook[sheet_name]

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    # Create a new sheet in the new workbook
    sheet_content = new_wb.create_sheet(title='content')

    # for CONTENT FILE create only content sheet
    if sheet_name == 'content':
        # Copy the data from the original sheet to the new sheet
        for row in original_sheet.iter_rows(values_only=True):
            sheet_content.append(row)
        # Remove the default sheet (Sheet)
        default_sheet = new_wb["Sheet"]
        new_wb.remove(default_sheet)
        # adjust cell width
        new_wb = adjust_cell_width(new_wb)
        # Save the workbook to a file
        new_wb.save(save_path + sheet_name + '.xlsx')
        # BREAK LOOP
        continue

    # FOR NON-CONTENT FILE
    key = ntoi[stoo[sheet_name]]
    continuous, factorial = describe(dfs[key])

    # create content sheet
    continuous_names = set([row[0] for row in continuous[1:]])
    sheet_content.append(['name', 'type'])
    for row in original_sheet.iter_rows(values_only=True):
        category = 'continuous' if row[0] in continuous_names else 'factorial'
        sheet_content.append(list(row) + [category])
    # Create a new sheet for description od continuous variables
    sheet_con = new_wb.create_sheet(title='continuous')
    for row in continuous:
        sheet_con.append(row)
    sheet_fact = new_wb.create_sheet(title='factorial')
    try:
        end_numeric = factorial[0].index('level_distribution')
    except:
        end_numeric = 4
    for row in factorial:
        # suppress invalid characters
        new_row = row[:end_numeric] + [ILLEGAL_CHARACTERS_RE.sub(r'', str(el)) for el in row[end_numeric:]]
        sheet_fact.append(new_row)
    sheet_fact_T = new_wb.create_sheet(title='factorial_transpose')
    # transpose factorial list
    max_len_factorial = max(len(row) for row in factorial)
    for row in factorial:
        while len(row) < max_len_factorial:
            row.append('')
    factorial = list(zip(*factorial))
    for index, row in enumerate(factorial):
        # suppress invalid characters
        if index >= end_numeric:
            row = [ILLEGAL_CHARACTERS_RE.sub(r'', str(el)) for el in row]
        sheet_fact_T.append(row)


    # Remove the default sheet (Sheet)
    default_sheet = new_wb["Sheet"]
    new_wb.remove(default_sheet)
    # adjust cell width
    new_wb = adjust_cell_width(new_wb)
    # Save the workbook to a file
    new_wb.save(save_path + str(ntoi[stoo[sheet_name]]).zfill(2) + '_' + stoo[sheet_name] + '.xlsx')
    print(f'sheet {str(ntoi[stoo[sheet_name]]).zfill(2)}_{sheet_name} done')


print('\nFINISHED')

